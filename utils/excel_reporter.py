"""
Excel 报告生成工具模块

生成横向格式的 Excel 报告，包含文件夹和图片 URL。
"""

from pathlib import Path
from typing import Dict, List
from utils.logger import log_info, log_error


def generate_excel_report(upload_results: Dict[str, List[Dict]], output_path: str) -> bool:
    """
    生成 Excel 报告

    Args:
        upload_results: 上传结果字典，格式为 {folder_name: [{'filename': ..., 'url': ...}, ...]}
        output_path: 输出文件路径

    Returns:
        bool: 成功返回 True，失败返回 False
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        log_error("[Excel] 请先安装 openpyxl 库: pip install openpyxl")
        return False

    try:
        # 贴建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "图片上传报告"

        # 设置标题样式
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # 写入表头
        ws.cell(1, 1, "文件夹名称").font = header_font
        ws.cell(1, 1).fill = header_fill
        ws.cell(1, 1).alignment = header_alignment

        # 找出最大图片数量
        max_images = max(len(images) for images in upload_results.values()) if upload_results else 0

        # 写入图片列标题
        for i in range(max_images):
            col = i + 2
            ws.cell(1, col, f"图片 {i + 1}").font = header_font
            ws.cell(1, col).fill = header_fill
            ws.cell(1, col).alignment = header_alignment

        # 写入数据
        row = 2
        for folder_name, images in sorted(upload_results.items()):
            ws.cell(row, 1, folder_name)

            for col_idx, image_info in enumerate(images):
                col = col_idx + 2
                url = image_info.get('url', '')
                filename = image_info.get('filename', '')

                # 写入 URL（带超链接）
                cell = ws.cell(row, col, url)
                if url:
                    cell.hyperlink = url
                    cell.font = Font(color="0563C1", underline="single")
                    cell.alignment = Alignment(horizontal="left", vertical="center")

            row += 1

        # 调整列宽
        ws.column_dimensions['A'].width = 30
        for i in range(max_images):
            col_letter = openpyxl.utils.get_column_letter(i + 2)
            ws.column_dimensions[col_letter].width = 60

        # 保存文件
        wb.save(output_path)
        log_info(f"[Excel] 报告生成成功: {output_path}")
        return True

    except Exception as e:
        log_error(f"[Excel] 生成报告失败: {e}")
        return False


def generate_simple_excel_report(image_urls: List[Dict], output_path: str) -> bool:
    """
    生成简单的 Excel 报告（单列格式）

    Args:
        image_urls: 图片 URL 列表，格式为 [{'filename': ..., 'url': ...}, ...]
        output_path: 输出文件路径

    Returns:
        bool: 成功返回 True，失败返回 False
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        log_error("[Excel] 请先安装 openpyxl 库: pip install openpyxl")
        return False

    try:
        # 贴建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "图片上传报告"

        # 设置标题样式
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # 写入表头
        ws.cell(1, 1, "文件名").font = header_font
        ws.cell(1, 1).fill = header_fill
        ws.cell(1, 1).alignment = header_alignment

        ws.cell(1, 2, "图片 URL").font = header_font
        ws.cell(1, 2).fill = header_fill
        ws.cell(1, 2).alignment = header_alignment

        # 写入数据
        for row, image_info in enumerate(image_urls, start=2):
            filename = image_info.get('filename', '')
            url = image_info.get('url', '')

            ws.cell(row, 1, filename)

            # 写入 URL（带超链接）
            cell = ws.cell(row, 2, url)
            if url:
                cell.hyperlink = url
                cell.font = Font(color="0563C1", underline="single")
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # 调整列宽
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 80

        # 保存文件
        wb.save(output_path)
        log_info(f"[Excel] 报告生成成功: {output_path}")
        return True

    except Exception as e:
        log_error(f"[Excel] 生成报告失败: {e}")
        return False
